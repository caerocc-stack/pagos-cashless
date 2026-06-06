"""
Backup automatico de la base de datos a un archivo Excel multi-hoja.
Se ejecuta desde GitHub Actions todos los dias.
Reusa la conexion y modelos de la aplicacion.
"""
import os
import sys
from pathlib import Path
from datetime import datetime, timezone, timedelta

# Permitir importar el paquete app
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import SessionLocal
from app.models.alumno import Alumno
from app.models.tarjeta import Tarjeta
from app.models.saldo import Saldo
from app.models.movimiento import Movimiento

AR_TZ = timezone(timedelta(hours=-3))


def encabezado(ws, columnas):
    fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for i, t in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")


def fecha(dt):
    return dt.strftime("%d/%m/%Y %H:%M") if dt else ""


def main():
    db = SessionLocal()
    try:
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Alumnos"
        encabezado(ws, ["id", "legajo", "dni", "apellido", "nombre", "curso", "activo", "created_at"])
        for a in db.query(Alumno).order_by(Alumno.id).all():
            ws.append([a.id, a.legajo, a.dni, a.apellido, a.nombre, a.curso,
                       "Si" if a.activo else "No", fecha(a.created_at)])

        ws = wb.create_sheet("Saldos")
        encabezado(ws, ["alumno_id", "monto", "updated_at"])
        for s in db.query(Saldo).order_by(Saldo.alumno_id).all():
            ws.append([s.alumno_id, float(s.monto), fecha(s.updated_at)])

        ws = wb.create_sheet("Tarjetas")
        encabezado(ws, ["id", "uid", "alumno_id", "activa", "created_at"])
        for t in db.query(Tarjeta).order_by(Tarjeta.id).all():
            ws.append([t.id, t.uid, t.alumno_id, "Si" if t.activa else "No", fecha(t.created_at)])

        ws = wb.create_sheet("Movimientos")
        encabezado(ws, ["id", "alumno_id", "tipo", "monto", "descripcion", "referencia_id", "operador", "created_at"])
        for m in db.query(Movimiento).order_by(Movimiento.id).all():
            ws.append([m.id, m.alumno_id, m.tipo, float(m.monto), m.descripcion or "",
                       m.referencia_id or "", m.operador or "", fecha(m.created_at)])

        for hoja in wb.worksheets:
            for col in hoja.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                hoja.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        carpeta = Path(__file__).resolve().parent.parent / "backups"
        carpeta.mkdir(exist_ok=True)
        nombre = f"backup_apai_{datetime.now(AR_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
        ruta = carpeta / nombre
        wb.save(ruta)
        print(f"Backup generado: {ruta}")
        print(f"Alumnos: {db.query(Alumno).count()}, Movimientos: {db.query(Movimiento).count()}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
