# -*- coding: utf-8 -*-
"""Importa los proveedores extraidos de las planillas de gastos a la base."""
import sys
import re
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from app.database import SessionLocal, engine, Base
from app.models.proveedor import Proveedor

ARCHIVOS = [
    r'C:\Users\Carlos\Desktop\Factura PDF\Gastos Marzo 2026.xlsx',
    r'C:\Users\Carlos\Downloads\Gastos MARZO-2026.xlsx',
]
HOJAS = ['Gastos de Oficina', 'Gtos. de Oficina', 'Infraestructura',
         'Gastos de la Escuela', 'Gtos. de la Escuela', 'Honorarios']


def norm_cuit(c):
    if c is None:
        return None
    s = ''.join(ch for ch in str(c) if ch.isdigit())
    return s if len(s) == 11 else None


def separar_fantasia(razon):
    """Separa nombre de fantasia si viene entre parentesis o tras ' - '."""
    razon = str(razon).strip()
    m = re.search(r'\(([^)]+)\)', razon)
    if m:
        fant = m.group(1).strip()
        base = razon[:m.start()].strip(' -')
        return base, fant
    if ' - ' in razon:
        base, fant = razon.split(' - ', 1)
        return base.strip(), fant.strip()
    return razon, None


# Asegurar que la tabla exista
Base.metadata.create_all(bind=engine, tables=[Proveedor.__table__])

provs = {}
for f in ARCHIVOS:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    for h in wb.sheetnames:
        if h not in HOJAS:
            continue
        for r in wb[h].iter_rows(min_row=4, values_only=True):
            if len(r) < 9:
                continue
            razon, cuit, rubro = r[4], norm_cuit(r[5]), r[8]
            if not cuit or not razon:
                continue
            if cuit not in provs:
                provs[cuit] = {'razon': str(razon).strip(), 'rubros': Counter()}
            if rubro:
                provs[cuit]['rubros'][str(rubro).strip()] += 1

db = SessionLocal()
existentes = {p.cuit for p in db.query(Proveedor.cuit).all()}
creados = 0
for cuit, p in provs.items():
    if cuit in existentes:
        continue
    base, fant = separar_fantasia(p['razon'])
    rubro = p['rubros'].most_common(1)[0][0] if p['rubros'] else None
    db.add(Proveedor(cuit=cuit, razon_social=base[:160], nombre_fantasia=(fant[:160] if fant else None),
                     rubro=(rubro[:80] if rubro else None)))
    creados += 1
db.commit()
print(f"Proveedores creados: {creados} | ya existian: {len(existentes)} | total en planilla: {len(provs)}")
db.close()
