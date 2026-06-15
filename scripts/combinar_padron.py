# -*- coding: utf-8 -*-
"""Combina el padron ALUMNOS 2026 (Excel) con la base del sistema.

- Cruza por legajo base (5 digitos = primeros 5 del legajo de 6 del Excel).
- Enriquece los alumnos existentes (division, modalidad, condicion, contacto, etc.).
- Da de alta los que faltan.
- Por defecto corre en modo PRUEBA (no escribe). Pasar --commit para aplicar.

Uso:  python scripts/combinar_padron.py            (prueba)
      python scripts/combinar_padron.py --commit   (aplica)
"""
import sys
import re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import app.main  # noqa: F401 (registra modelos / migraciones)
from app.database import SessionLocal
from app.models.alumno import Alumno

ARCHIVO = r'C:\Users\Carlos\Downloads\ALUMNOS Y HORARIOS (2).xlsx'
HOJA = 'ALUMNOS 2026'
FILA_DATOS = 4
COMMIT = '--commit' in sys.argv


def dig(x):
    # Los numeros vienen como float (25199.0); pasar a int para no arrastrar el '0' del '.0'
    if isinstance(x, float) and x.is_integer():
        x = int(x)
    return "".join(ch for ch in str(x) if ch.isdigit()) if x is not None else ""


def legajo_base(x):
    return dig(x)  # los legajos son de 5 digitos, sin verificador


def partir_nombre(completo):
    """'DA COSTA CARDANTE David' -> ('DA COSTA CARDANTE', 'David'). Apellido = prefijo en MAYUSCULAS."""
    txt = str(completo).replace(",", " ").strip()
    palabras = txt.split()
    ape = []
    i = 0
    while i < len(palabras) and palabras[i].isupper() and any(c.isalpha() for c in palabras[i]):
        ape.append(palabras[i])
        i += 1
    if not ape:  # no hay prefijo en mayusculas: primera palabra = apellido
        ape = palabras[:1]
        i = 1
    apellido = " ".join(ape)
    nombre = " ".join(palabras[i:]) or apellido
    return apellido[:100], nombre[:100]


def derivar_division(curso):
    c = str(curso or "").strip()
    if len(c) >= 2 and c[-1].isalpha() and (c[-2] in "°º -/" or c[-2].isdigit()):
        return c[-1].upper()
    return None


def main():
    wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
    ws = wb[HOJA]
    db = SessionLocal()

    # Indice de existentes por legajo base
    existentes = {dig(a.legajo): a for a in db.query(Alumno).all() if a.legajo}
    dnis_usados = {a.dni for a in db.query(Alumno).all() if a.dni}

    nuevos = enriquecidos = saltados = 0
    for r in range(FILA_DATOS, ws.max_row + 1):
        nom = ws.cell(r, 2).value
        if not nom:
            continue
        base = legajo_base(ws.cell(r, 4).value)
        if not base:
            continue
        leg_comp = dig(ws.cell(r, 4).value)
        dni = dig(ws.cell(r, 3).value)[:10] or None
        area = ws.cell(r, 5).value
        curso = ws.cell(r, 6).value
        condicion = ws.cell(r, 32).value      # AF
        modalidad = ws.cell(r, 33).value      # AG
        fnac = ws.cell(r, 59).value           # BG
        tel = dig(ws.cell(r, 60).value) or None  # BH
        email = ws.cell(r, 61).value          # BI
        division = derivar_division(curso)
        apellido, nombre = partir_nombre(nom)

        a = existentes.get(base)
        if a:
            # Enriquecer (sin pisar legajo/nombre que ya estan bien)
            a.area = str(area) if area else a.area
            a.curso = str(curso) if curso else a.curso
            a.division = division or a.division
            a.condicion = str(condicion) if condicion else a.condicion
            a.modalidad = str(modalidad) if modalidad else a.modalidad
            a.telefono = tel or a.telefono
            a.email = (str(email) if email else None) or a.email
            a.legajo_completo = leg_comp or a.legajo_completo
            if hasattr(fnac, "date"):
                a.fecha_nacimiento = fnac.date()
            # DNI: solo si difiere y no lo usa otro alumno
            if dni and dni != a.dni and dni not in dnis_usados:
                dnis_usados.discard(a.dni)
                a.dni = dni
                dnis_usados.add(dni)
            enriquecidos += 1
        else:
            # Alta nueva. DNI obligatorio y unico en el modelo.
            if not dni or dni in dnis_usados:
                dni = f"SL{base}"[:10]  # sin DNI valido: placeholder unico
            nuevo = Alumno(
                legajo=base, dni=dni, nombre=nombre, apellido=apellido,
                curso=str(curso) if curso else "", area=str(area) if area else None,
                division=division, condicion=str(condicion) if condicion else None,
                modalidad=str(modalidad) if modalidad else None,
                telefono=tel, email=str(email) if email else None,
                legajo_completo=leg_comp,
                fecha_nacimiento=fnac.date() if hasattr(fnac, "date") else None,
            )
            db.add(nuevo)
            dnis_usados.add(dni)
            existentes[base] = nuevo
            nuevos += 1

    print(f"Enriquecidos: {enriquecidos} | Nuevos: {nuevos} | Saltados: {saltados}")
    if COMMIT:
        db.commit()
        print(">>> CAMBIOS APLICADOS A LA BASE <<<")
    else:
        db.rollback()
        print("(modo prueba: no se escribio nada. Correr con --commit para aplicar)")
    db.close()


if __name__ == "__main__":
    main()
