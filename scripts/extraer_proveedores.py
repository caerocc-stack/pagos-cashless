# -*- coding: utf-8 -*-
import openpyxl
from collections import Counter


def norm_cuit(c):
    if c is None:
        return None
    s = ''.join(ch for ch in str(c) if ch.isdigit())
    return s if len(s) == 11 else None


provs = {}
archivos = [
    r'C:\Users\Carlos\Desktop\Factura PDF\Gastos Marzo 2026.xlsx',
    r'C:\Users\Carlos\Downloads\Gastos MARZO-2026.xlsx',
]
hojas_gasto = ['Gastos de Oficina', 'Gtos. de Oficina', 'Infraestructura',
               'Gastos de la Escuela', 'Gtos. de la Escuela', 'Honorarios']

for f in archivos:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    for h in wb.sheetnames:
        if h not in hojas_gasto:
            continue
        ws = wb[h]
        for r in ws.iter_rows(min_row=4, values_only=True):
            if len(r) < 9:
                continue
            razon, cuit, rubro = r[4], norm_cuit(r[5]), r[8]
            if not cuit or not razon:
                continue
            razon = str(razon).strip()
            if cuit not in provs:
                provs[cuit] = {'razon': razon, 'rubros': Counter(), 'n': 0}
            if rubro:
                provs[cuit]['rubros'][str(rubro).strip()] += 1
            provs[cuit]['n'] += 1


def sa(s):
    return str(s).encode('ascii', 'replace').decode()


print('PROVEEDORES UNICOS DETECTADOS:', len(provs))
print('Total facturas analizadas:', sum(p['n'] for p in provs.values()))
print('--- Muestra (20 con mas facturas) ---')
for cuit, p in sorted(provs.items(), key=lambda x: -x[1]['n'])[:20]:
    rubro = p['rubros'].most_common(1)[0][0] if p['rubros'] else '-'
    print(sa('  {}  {:34}  rubro: {:22}  ({} fact.)'.format(
        cuit, p['razon'][:34], rubro[:22], p['n'])))
