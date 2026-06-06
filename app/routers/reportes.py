from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case
from sqlalchemy.orm import Session
from datetime import datetime, timedelta, date
from decimal import Decimal
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.auth import get_current_user
from app.models.alumno import Alumno
from app.models.saldo import Saldo
from app.models.movimiento import Movimiento
from app.tz import ahora_ar

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])


def _rango_fechas(desde: str | None, hasta: str | None):
    """Convierte strings YYYY-MM-DD a datetime. Si no hay, usa ultimos 30 dias."""
    if hasta:
        dt_hasta = datetime.strptime(hasta, "%Y-%m-%d") + timedelta(days=1)
    else:
        dt_hasta = ahora_ar() + timedelta(days=1)
    if desde:
        dt_desde = datetime.strptime(desde, "%Y-%m-%d")
    else:
        dt_desde = (ahora_ar() - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
    return dt_desde, dt_hasta


@router.get("/resumen")
def resumen_general(
    desde: str | None = None,
    hasta: str | None = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Totales y cantidades de operaciones en un rango de fechas."""
    dt_desde, dt_hasta = _rango_fechas(desde, hasta)

    base = db.query(Movimiento).filter(
        Movimiento.created_at >= dt_desde,
        Movimiento.created_at < dt_hasta,
    )

    def totales(tipo):
        q = base.filter(Movimiento.tipo == tipo)
        suma = db.query(func.coalesce(func.sum(func.abs(Movimiento.monto)), 0)).filter(
            Movimiento.tipo == tipo,
            Movimiento.created_at >= dt_desde,
            Movimiento.created_at < dt_hasta,
        ).scalar()
        cantidad = q.count()
        return float(suma or 0), cantidad

    recargas_monto, recargas_cant = totales("recarga")
    consumos_monto, consumos_cant = totales("consumo")
    reintegros_monto, reintegros_cant = totales("reintegro")
    transf_monto, transf_cant = totales("transferencia_out")

    # Saldo total actual en el sistema
    saldo_total = db.query(func.coalesce(func.sum(Saldo.monto), 0)).scalar()

    total_alumnos = db.query(func.count(Alumno.id)).scalar()
    alumnos_con_saldo = db.query(func.count(Saldo.alumno_id)).filter(Saldo.monto > 0).scalar()

    promedio_consumo = consumos_monto / consumos_cant if consumos_cant else 0

    return {
        "desde": dt_desde.strftime("%Y-%m-%d"),
        "hasta": (dt_hasta - timedelta(days=1)).strftime("%Y-%m-%d"),
        "recargas": {"monto": recargas_monto, "cantidad": recargas_cant},
        "consumos": {"monto": consumos_monto, "cantidad": consumos_cant},
        "reintegros": {"monto": reintegros_monto, "cantidad": reintegros_cant},
        "transferencias": {"monto": transf_monto, "cantidad": transf_cant},
        "saldo_total_sistema": float(saldo_total or 0),
        "total_alumnos": total_alumnos,
        "alumnos_con_saldo": alumnos_con_saldo,
        "promedio_consumo": round(promedio_consumo, 2),
        "ticket_promedio": round(promedio_consumo, 2),
    }


@router.get("/por-curso")
def consumo_por_curso(
    desde: str | None = None,
    hasta: str | None = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Consumo total y promedio por curso."""
    dt_desde, dt_hasta = _rango_fechas(desde, hasta)

    # Consumo agregado por curso (solo movimientos tipo consumo)
    filas = (
        db.query(
            Alumno.curso.label("curso"),
            func.coalesce(func.sum(func.abs(Movimiento.monto)), 0).label("total"),
            func.count(Movimiento.id).label("cantidad"),
            func.count(func.distinct(Movimiento.alumno_id)).label("alumnos_compradores"),
        )
        .join(Movimiento, Movimiento.alumno_id == Alumno.id)
        .filter(
            Movimiento.tipo == "consumo",
            Movimiento.created_at >= dt_desde,
            Movimiento.created_at < dt_hasta,
        )
        .group_by(Alumno.curso)
        .all()
    )

    # Cantidad de alumnos por curso (para promedio real)
    alumnos_por_curso = dict(
        db.query(Alumno.curso, func.count(Alumno.id)).group_by(Alumno.curso).all()
    )

    resultado = []
    for f in filas:
        total = float(f.total or 0)
        total_alumnos_curso = alumnos_por_curso.get(f.curso, 0) or 1
        resultado.append({
            "curso": f.curso,
            "total_consumo": total,
            "cantidad_operaciones": f.cantidad,
            "alumnos_compradores": f.alumnos_compradores,
            "alumnos_curso": alumnos_por_curso.get(f.curso, 0),
            "promedio_por_alumno": round(total / total_alumnos_curso, 2),
            "promedio_por_operacion": round(total / f.cantidad, 2) if f.cantidad else 0,
        })

    resultado.sort(key=lambda x: x["total_consumo"], reverse=True)
    return resultado


@router.get("/diario")
def consumo_diario(
    dias: int = 30,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Consumo total por dia para grafico de tendencia."""
    desde = (ahora_ar() - timedelta(days=dias)).replace(hour=0, minute=0, second=0, microsecond=0)

    filas = (
        db.query(
            func.date(Movimiento.created_at).label("dia"),
            func.coalesce(func.sum(func.abs(Movimiento.monto)), 0).label("total"),
            func.count(Movimiento.id).label("cantidad"),
        )
        .filter(
            Movimiento.tipo == "consumo",
            Movimiento.created_at >= desde,
        )
        .group_by(func.date(Movimiento.created_at))
        .order_by(func.date(Movimiento.created_at))
        .all()
    )

    return [
        {"dia": str(f.dia), "total": float(f.total or 0), "cantidad": f.cantidad}
        for f in filas
    ]


@router.get("/top-alumnos")
def top_alumnos(
    limite: int = 10,
    desde: str | None = None,
    hasta: str | None = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Alumnos que mas consumen en el rango."""
    dt_desde, dt_hasta = _rango_fechas(desde, hasta)

    filas = (
        db.query(
            Alumno.id,
            Alumno.apellido,
            Alumno.nombre,
            Alumno.curso,
            Alumno.legajo,
            func.coalesce(func.sum(func.abs(Movimiento.monto)), 0).label("total"),
            func.count(Movimiento.id).label("cantidad"),
        )
        .join(Movimiento, Movimiento.alumno_id == Alumno.id)
        .filter(
            Movimiento.tipo == "consumo",
            Movimiento.created_at >= dt_desde,
            Movimiento.created_at < dt_hasta,
        )
        .group_by(Alumno.id, Alumno.apellido, Alumno.nombre, Alumno.curso, Alumno.legajo)
        .order_by(func.sum(func.abs(Movimiento.monto)).desc())
        .limit(limite)
        .all()
    )

    return [
        {
            "alumno_id": f.id,
            "apellido": f.apellido,
            "nombre": f.nombre,
            "curso": f.curso,
            "legajo": f.legajo,
            "total_consumo": float(f.total or 0),
            "cantidad": f.cantidad,
        }
        for f in filas
    ]


def _estilo_encabezado(ws, columnas):
    """Aplica estilo al encabezado de una hoja Excel."""
    fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col_num, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=titulo)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


@router.get("/exportar-excel")
def exportar_excel(
    tipo: str = Query("movimientos", description="movimientos | alumnos | resumen | por-curso"),
    desde: str | None = None,
    hasta: str | None = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Genera un archivo Excel descargable segun el tipo de reporte."""
    wb = openpyxl.Workbook()
    ws = wb.active
    fecha_str = ahora_ar().strftime("%Y%m%d_%H%M")

    if tipo == "alumnos":
        ws.title = "Alumnos"
        _estilo_encabezado(ws, ["Legajo", "DNI", "Apellido", "Nombre", "Curso", "Saldo"])
        from sqlalchemy.orm import joinedload
        alumnos = (
            db.query(Alumno).options(joinedload(Alumno.saldo))
            .order_by(Alumno.apellido, Alumno.nombre).all()
        )
        for a in alumnos:
            saldo = float(a.saldo.monto) if a.saldo else 0
            ws.append([a.legajo, a.dni, a.apellido, a.nombre, a.curso, saldo])
        nombre_archivo = f"alumnos_{fecha_str}.xlsx"

    elif tipo == "por-curso":
        ws.title = "Consumo por curso"
        _estilo_encabezado(ws, ["Curso", "Total consumo", "Operaciones", "Alumnos compradores", "Alumnos curso", "Promedio x alumno"])
        datos = consumo_por_curso(desde, hasta, db, user)
        for d in datos:
            ws.append([
                d["curso"], d["total_consumo"], d["cantidad_operaciones"],
                d["alumnos_compradores"], d["alumnos_curso"], d["promedio_por_alumno"],
            ])
        nombre_archivo = f"consumo_por_curso_{fecha_str}.xlsx"

    elif tipo == "resumen":
        ws.title = "Resumen"
        r = resumen_general(desde, hasta, db, user)
        _estilo_encabezado(ws, ["Concepto", "Monto", "Cantidad"])
        ws.append(["Recargas", r["recargas"]["monto"], r["recargas"]["cantidad"]])
        ws.append(["Consumos", r["consumos"]["monto"], r["consumos"]["cantidad"]])
        ws.append(["Reintegros", r["reintegros"]["monto"], r["reintegros"]["cantidad"]])
        ws.append(["Transferencias", r["transferencias"]["monto"], r["transferencias"]["cantidad"]])
        ws.append([])
        ws.append(["Saldo total en sistema", r["saldo_total_sistema"], ""])
        ws.append(["Total alumnos", r["total_alumnos"], ""])
        ws.append(["Alumnos con saldo", r["alumnos_con_saldo"], ""])
        ws.append(["Ticket promedio consumo", r["ticket_promedio"], ""])
        nombre_archivo = f"resumen_{fecha_str}.xlsx"

    else:  # movimientos
        ws.title = "Movimientos"
        _estilo_encabezado(ws, ["Fecha", "Legajo", "Apellido", "Nombre", "Curso", "Tipo", "Monto", "Descripcion", "Operador"])
        dt_desde, dt_hasta = _rango_fechas(desde, hasta)
        movs = (
            db.query(Movimiento, Alumno)
            .join(Alumno, Movimiento.alumno_id == Alumno.id)
            .filter(Movimiento.created_at >= dt_desde, Movimiento.created_at < dt_hasta)
            .order_by(Movimiento.created_at.desc())
            .all()
        )
        for m, a in movs:
            ws.append([
                m.created_at.strftime("%d/%m/%Y %H:%M"),
                a.legajo, a.apellido, a.nombre, a.curso,
                m.tipo, float(m.monto), m.descripcion or "", m.operador or "",
            ])
        nombre_archivo = f"movimientos_{fecha_str}.xlsx"

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )
