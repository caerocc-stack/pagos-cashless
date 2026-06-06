from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.auth import get_current_user
from app.models.alumno import Alumno
from app.models.tarjeta import Tarjeta
from app.models.saldo import Saldo
from app.models.movimiento import Movimiento
from app.tz import ahora_ar
from app import email_util

router = APIRouter(prefix="/api/admin", tags=["Administracion"])


class TestEmailRequest(BaseModel):
    destino: str | None = None


@router.post("/test-email")
def test_email(data: TestEmailRequest, user=Depends(get_current_user)):
    """Envia un email de prueba para verificar la configuracion SMTP."""
    if not email_util.email_configurado():
        raise HTTPException(
            400,
            "El envío de email no está configurado todavía. Cargá SMTP_USER y SMTP_PASS "
            "en las variables de entorno de Render.",
        )
    destino = (data.destino or "").strip() or os.getenv("SMTP_USER")
    cuerpo = """
    <div style="font-family:Segoe UI,Arial,sans-serif;max-width:480px;margin:auto;color:#15233b">
        <div style="background:#15233b;color:#fff;padding:16px 20px;border-radius:10px 10px 0 0">
            <h2 style="margin:0">APAI Pay</h2>
        </div>
        <div style="border:1px solid #e2e8f0;border-top:none;padding:20px;border-radius:0 0 10px 10px">
            <p>¡El envío de correos quedó configurado correctamente! ✅</p>
            <p style="font-size:13px;color:#5b6b80">Este es un email de prueba del sistema APAI Pay.</p>
        </div>
    </div>"""
    try:
        email_util.enviar_email(destino, "Prueba de configuración - APAI Pay", cuerpo)
    except Exception as e:
        raise HTTPException(502, f"No se pudo enviar el email: {e}")
    return {"ok": True, "mensaje": f"Email de prueba enviado a {destino}"}


def _encabezado(ws, columnas):
    fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for i, t in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")


@router.get("/backup")
def backup_completo(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """
    Genera un backup completo de toda la base de datos en un Excel
    con una hoja por cada tabla (alumnos, tarjetas, saldos, movimientos).
    """
    wb = openpyxl.Workbook()

    # Hoja Alumnos
    ws = wb.active
    ws.title = "Alumnos"
    _encabezado(ws, ["id", "legajo", "dni", "apellido", "nombre", "curso", "activo", "created_at"])
    for a in db.query(Alumno).order_by(Alumno.id).all():
        ws.append([
            a.id, a.legajo, a.dni, a.apellido, a.nombre, a.curso,
            "Si" if a.activo else "No",
            a.created_at.strftime("%d/%m/%Y %H:%M") if a.created_at else "",
        ])

    # Hoja Saldos
    ws = wb.create_sheet("Saldos")
    _encabezado(ws, ["alumno_id", "monto", "updated_at"])
    for s in db.query(Saldo).order_by(Saldo.alumno_id).all():
        ws.append([
            s.alumno_id, float(s.monto),
            s.updated_at.strftime("%d/%m/%Y %H:%M") if s.updated_at else "",
        ])

    # Hoja Tarjetas
    ws = wb.create_sheet("Tarjetas")
    _encabezado(ws, ["id", "uid", "alumno_id", "activa", "created_at"])
    for t in db.query(Tarjeta).order_by(Tarjeta.id).all():
        ws.append([
            t.id, t.uid, t.alumno_id,
            "Si" if t.activa else "No",
            t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else "",
        ])

    # Hoja Movimientos
    ws = wb.create_sheet("Movimientos")
    _encabezado(ws, ["id", "alumno_id", "tipo", "monto", "descripcion", "referencia_id", "operador", "created_at"])
    for m in db.query(Movimiento).order_by(Movimiento.id).all():
        ws.append([
            m.id, m.alumno_id, m.tipo, float(m.monto),
            m.descripcion or "", m.referencia_id or "", m.operador or "",
            m.created_at.strftime("%d/%m/%Y %H:%M") if m.created_at else "",
        ])

    # Ajustar anchos
    for hoja in wb.worksheets:
        for col in hoja.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            hoja.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha = ahora_ar().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=backup_apai_{fecha}.xlsx"},
    )
