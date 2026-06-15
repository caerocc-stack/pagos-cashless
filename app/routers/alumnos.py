from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from decimal import Decimal
import openpyxl
import io

from app.database import get_db
from app.models.alumno import Alumno
from app.models.saldo import Saldo
from app.models.tarjeta import Tarjeta
from app.models.movimiento import Movimiento
from app.schemas.alumno import AlumnoCreate, AlumnoUpdate, AlumnoResponse, AlumnoConSaldo

router = APIRouter(prefix="/api/alumnos", tags=["Alumnos"])


def derivar_area(curso: str) -> str | None:
    """Deduce el area a partir del nombre del curso."""
    c = (curso or "").lower()
    if "avionica" in c or "aviónica" in c:
        return "Avionica"
    if "mecanica" in c or "mecánica" in c:
        return "Mecanica"
    if "ciclo" in c or "basico" in c or "básico" in c:
        return "Ciclo Basico"
    return None


def _alumno_con_saldo(a: Alumno) -> dict:
    return {
        "id": a.id,
        "legajo": a.legajo,
        "dni": a.dni,
        "nombre": a.nombre,
        "apellido": a.apellido,
        "curso": a.curso,
        "area": a.area or derivar_area(a.curso),
        "division": a.division,
        "email": a.email,
        "telefono": a.telefono,
        "fecha_nacimiento": a.fecha_nacimiento,
        "condicion": a.condicion,
        "modalidad": a.modalidad,
        "cuota_excluir": a.cuota_excluir,
        "cuota_personalizada": a.cuota_personalizada,
        "activo": a.activo,
        "created_at": a.created_at,
        "saldo": a.saldo.monto if a.saldo else Decimal("0.00"),
        "codigo_siro": a.codigo_siro,
    }


@router.get("/", response_model=list[AlumnoConSaldo])
def listar_alumnos(activo: bool | None = None, curso: str | None = None, db: Session = Depends(get_db)):
    query = db.query(Alumno).options(joinedload(Alumno.saldo))
    if activo is not None:
        query = query.filter(Alumno.activo == activo)
    if curso:
        query = query.filter(Alumno.curso == curso)
    alumnos = query.order_by(Alumno.apellido, Alumno.nombre).all()
    return [_alumno_con_saldo(a) for a in alumnos]


@router.get("/{alumno_id}", response_model=AlumnoConSaldo)
def obtener_alumno(alumno_id: int, db: Session = Depends(get_db)):
    alumno = db.query(Alumno).options(joinedload(Alumno.saldo)).filter(Alumno.id == alumno_id).first()
    if not alumno:
        raise HTTPException(404, "Alumno no encontrado")
    return _alumno_con_saldo(alumno)


@router.post("/", response_model=AlumnoResponse, status_code=201)
def crear_alumno(data: AlumnoCreate, db: Session = Depends(get_db)):
    if db.query(Alumno).filter(Alumno.legajo == data.legajo).first():
        raise HTTPException(409, f"Ya existe un alumno con legajo {data.legajo}")
    if db.query(Alumno).filter(Alumno.dni == data.dni).first():
        raise HTTPException(409, f"Ya existe un alumno con DNI {data.dni}")

    alumno = Alumno(**data.model_dump())
    if not alumno.area:
        alumno.area = derivar_area(alumno.curso)
    db.add(alumno)
    db.flush()

    saldo = Saldo(alumno_id=alumno.id)
    db.add(saldo)
    db.commit()
    db.refresh(alumno)
    return alumno


@router.put("/{alumno_id}", response_model=AlumnoResponse)
def actualizar_alumno(alumno_id: int, data: AlumnoUpdate, db: Session = Depends(get_db)):
    alumno = db.get(Alumno, alumno_id)
    if not alumno:
        raise HTTPException(404, "Alumno no encontrado")
    for campo, valor in data.model_dump(exclude_unset=True).items():
        setattr(alumno, campo, valor)
    db.commit()
    db.refresh(alumno)
    return alumno


@router.delete("/eliminar-todos")
def eliminar_todos(db: Session = Depends(get_db)):
    """Elimina TODO el padron: alumnos, tarjetas, saldos y movimientos. Accion irreversible."""
    db.query(Movimiento).delete()
    db.query(Tarjeta).delete()
    db.query(Saldo).delete()
    cantidad = db.query(Alumno).delete()
    db.commit()
    return {"ok": True, "mensaje": f"Se eliminaron {cantidad} alumnos y todos sus datos asociados"}


@router.delete("/{alumno_id}")
def eliminar_alumno(alumno_id: int, db: Session = Depends(get_db)):
    """Elimina un alumno y todos sus datos asociados (tarjetas, saldo, movimientos)."""
    alumno = db.get(Alumno, alumno_id)
    if not alumno:
        raise HTTPException(404, "Alumno no encontrado")

    # Borrar movimientos, tarjetas y saldo del alumno
    db.query(Movimiento).filter(Movimiento.alumno_id == alumno_id).delete()
    db.query(Tarjeta).filter(Tarjeta.alumno_id == alumno_id).delete()
    db.query(Saldo).filter(Saldo.alumno_id == alumno_id).delete()
    db.delete(alumno)
    db.commit()

    return {"ok": True, "mensaje": f"Alumno {alumno.apellido}, {alumno.nombre} eliminado"}


@router.post("/importar")
def importar_excel(archivo: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Importa alumnos desde un archivo Excel.
    Columnas esperadas: legajo, dni, nombre, apellido, curso
    """
    contenido = archivo.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb.active

    encabezados = [str(cell.value).strip().lower() for cell in ws[1]]
    campos_requeridos = {"legajo", "dni", "nombre", "apellido", "curso"}
    faltantes = campos_requeridos - set(encabezados)
    if faltantes:
        raise HTTPException(400, f"Faltan columnas en el Excel: {', '.join(faltantes)}")

    col_map = {nombre: idx for idx, nombre in enumerate(encabezados)}
    # Columnas opcionales
    col_mail = col_map.get("mail", col_map.get("email"))
    col_area = col_map.get("area", col_map.get("área"))

    def _valor(fila, idx):
        if idx is None:
            return ""
        v = fila[idx]
        return "" if v is None else str(v).strip()

    # Cargar alumnos existentes (por legajo y por dni) para detectar duplicados y actualizar email
    existentes_por_legajo = {a.legajo: a for a in db.query(Alumno).all()}
    dnis_existentes = {a.dni for a in existentes_por_legajo.values()}

    creados = 0
    actualizados = 0
    errores = []
    nuevos_alumnos = []

    for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            legajo = _valor(fila, col_map["legajo"])
            dni = _valor(fila, col_map["dni"])
            nombre = _valor(fila, col_map["nombre"])
            apellido = _valor(fila, col_map["apellido"])
            curso = _valor(fila, col_map["curso"])
            email = _valor(fila, col_mail) or None
            area = _valor(fila, col_area) or derivar_area(curso)

            if not all([legajo, dni, nombre, apellido, curso]):
                errores.append(f"Fila {fila_num}: datos incompletos")
                continue

            # Si el alumno ya existe: actualizar email y area si el archivo los trae
            if legajo in existentes_por_legajo:
                ex = existentes_por_legajo[legajo]
                cambio = False
                if email:
                    ex.email = email
                    cambio = True
                if area and not ex.area:
                    ex.area = area
                    cambio = True
                if cambio:
                    actualizados += 1
                continue
            if dni in dnis_existentes:
                errores.append(f"Fila {fila_num}: DNI {dni} ya existe")
                continue

            nuevo = Alumno(
                legajo=legajo, dni=dni, nombre=nombre,
                apellido=apellido, curso=curso, email=email, area=area,
            )
            nuevos_alumnos.append(nuevo)
            existentes_por_legajo[legajo] = nuevo
            dnis_existentes.add(dni)
            creados += 1
        except Exception as e:
            errores.append(f"Fila {fila_num}: {str(e)}")

    # Insertar nuevos de una vez
    if nuevos_alumnos:
        db.add_all(nuevos_alumnos)
        db.flush()
        saldos = [Saldo(alumno_id=a.id) for a in nuevos_alumnos]
        db.add_all(saldos)

    db.commit()

    return {
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
        "total_procesadas": fila_num - 1 if 'fila_num' in dir() else 0,
    }
